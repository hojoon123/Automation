import openpyxl

save_path = '03.엑셀 자동화/11번가.xlsx'

wb = openpyxl.load_workbook(save_path, data_only = True)

ws = wb['data']

# 모든 셀의 데이터 가져오기
# -> 행과 열의 개수를 아는 경우
# for x in range(1, 10 + 1):
#     for y in range(1, 5 + 1):
#         print(ws.cell(row = x, column = y).value, end = " ")
#     print()
    
# -> 행과 열의 개수를 모르는 경우
# for x in range(1, ws.max_row + 1):
#     for y in range(1, ws.max_column + 1):
#         print(ws.cell(row = x, column = y).value, end = " ")
#     print()
    
# 모든 행 가져오기
# for row in ws.iter_rows():
#     print(row)

# 2번째 행부터 가져오기
# for row in ws.iter_rows(min_row = 2):
#     print(row)

# 2번쨰 행부터 5번째 행까지 가져오기
# for row in ws.iter_rows(min_row = 2, max_row = 5):
#     print(row)

# 2-4 행 2-4 열 가져오기
for row in ws.iter_rows(min_row = 2, min_col = 2, max_col = 4):
    for cell in row:
        print(cell.value, end = ' ')
    print()