import openpyxl

save_path = '03.엑셀 자동화/결혼정보 22.xlsx'

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path)

# 활성화 된 시트 선택
ws = wb.active

# 데이터 추가 (1)
ws['G1'] = '홈페이지'
ws['H1'] = '이메일'

# 데이터 추가 (2)
ws.cell(row = 2, column = 7, value = 'https://www.naver.com')
ws.cell(row = 2, column = 8, value = 'rhzn5512@naver.com')

# 데이터 추가 (3)
ws.append(['https://www.daum.net', 'rhzn5512@gmail.com'])


# 엑셀 저장
wb.save(save_path)