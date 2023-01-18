from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import chromedriver_autoinstaller as AutoChrome
import time
import pyperclip
import pyautogui
import openpyxl
import pandas as pd
import numpy as np

def start():
    global driver
    #크롬드라이버 버전 확인
    chrome_ver = AutoChrome.get_chrome_version().split('.')[0] 
    
    options = webdriver.ChromeOptions() # 브라우저 셋팅
    options.add_experimental_option("detach", True) # 브라우저 꺼짐 방지
    options.add_argument('lang=ko_KR') # 사용언어 한국어
    options.add_argument('disable-gpu') # 하드웨어 가속 안함
    options.add_experimental_option("excludeSwitches",['enable-logging']) # 불필요한 에러 메세지 삭제
    
    #실행 후 최신 버젼이 아니라서 실행이 안된다면 최신버젼으로 업데이트 후 재실행
    try:
        driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe', options = options)   
    except:
        AutoChrome.install(True)
        driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe', options = options)

    driver.implicitly_wait(10)
    
def openXlsx(save_path):
    global ws
    global wb
    # 기존 엑셀 파일 불러오기
    wb = openpyxl.load_workbook(save_path)
    # 활성화 된 시트 선택
    ws = wb.active
    # 데이터 추가
    ws['G1'] = '홈페이지'
    ws['H1'] = '이메일'
    
def loadData(save_pa):
    clip_data = pd.read_excel(save_pa, usecols = [2, 4])
    return clip_data
            

def naverMapOpen():
    # 네이버 로그인 주소 가져오기
    driver.get("https://map.naver.com")
    # 화면 최대화
    driver.maximize_window()
    
    driver.implicitly_wait(10)

def search():
    # 검색
    driver.find_element(By.CSS_SELECTOR, '#input_search1673979394698').click()
    driver.implicitly_wait(10)
    pyautogui.hotkey('ctrl', 'v')
    driver.implicitly_wait(3)
    pyautogui.press('enter')
    driver.implicitly_wait(10)


    
if __name__ == '__main__':
    sp = 'DB추출프로그램/결혼정보 22.xlsx'
    df = loadData(sp)
    for i in range(1,len(df)):
        print(df.loc[i])
        
    #start()
    #naverMapOpen()
    openXlsx(sp)
    #search()


   